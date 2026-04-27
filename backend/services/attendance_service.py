import io
import os
from datetime import datetime, timezone
from typing import Dict, Optional, List, Any

from fastapi import HTTPException, status, BackgroundTasks, logger, Path
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook
from sqlalchemy import select, insert, distinct
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models import Attendance, ModuleRegistration, AttendanceRequest, Student, Iam, Exam, Module, Lecturer
from backend.models.enums import RequestStatus, AttendanceStatus
from backend.schemas.attendance import AttendanceResponse, AttendanceCreate, AttendanceUpdate, AttendanceRequestCreate, \
    AttendanceRequestResponse, CorrectionApproval
from backend.services.email_service import send_attendance_request_confirmation, send_attendance_correction_notification
from backend.services.notification_service import create_new_request_notification, create_request_processed_notification

PRESENT_STATUSES = {AttendanceStatus.PRESENT, AttendanceStatus.LATE, AttendanceStatus.EXCUSED}
ELIGIBILITY_THRESHOLD = 80.0
EXPORT_DIR_NAME = "eligibility for exam"


async def get_attendance_by_id(session: AsyncSession, attendance_id: int) -> AttendanceResponse:
    result = await session.execute(select(Attendance).where(Attendance.attendance_id == attendance_id))
    attendance = result.scalars().first()
    if not attendance:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attendance not found")
    return attendance


async def submit_attendance(session: AsyncSession, data: AttendanceCreate) -> AttendanceResponse:
    reg_stmt = select(ModuleRegistration).where(
        ModuleRegistration.student_id == data.student_id,
        ModuleRegistration.module_id == data.module_id,
    )

    reg= await session.execute(reg_stmt)
    registration = reg.scalars().first()
    if not registration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Student {data.student_id} is not registered for module {data.module_id}")

    new_att = Attendance(**data.model_dump())
    session.add(new_att)
    await session.commit()
    await session.refresh(new_att)
    return AttendanceResponse.model_validate(new_att)


async def view_by_module(session: AsyncSession, module_id: str) -> list[AttendanceResponse]:
    result = await session.execute(select(Attendance).where(Attendance.module_id == module_id))
    attendances = result.scalars().all()
    return [AttendanceResponse.model_validate(attendance) for attendance in attendances]


async def view_by_student(session: AsyncSession, student_id: str) -> list[AttendanceResponse]:
    result = await session.execute(select(Attendance).where(Attendance.student_id == student_id))
    attendances = result.scalars().all()
    return [AttendanceResponse.model_validate(attendance) for attendance in attendances]


async def update_attendance(session: AsyncSession, attendance_id: int, data: AttendanceUpdate) -> AttendanceResponse:
    att = await session.get(Attendance, attendance_id)
    if not att:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attendance not found")
    if data.attendance_status:
        att.attendance_status = data.attendance_status
    await session.commit()
    await session.refresh(att)
    return AttendanceResponse.model_validate(att)


async def delete_attendance(session: AsyncSession, attendance_id: int) -> None:
    att = await session.get(Attendance, attendance_id)
    if not att:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attendance not found")
    await session.delete(att)


async def view_attendance_requests_by_module(
    session: AsyncSession, module_id: str
) -> List[AttendanceRequestResponse]:
    stmt = select(AttendanceRequest).where(AttendanceRequest.module_id == module_id)
    rows = (await session.execute(stmt)).scalars().all()
    return [AttendanceRequestResponse.model_validate(r) for r in rows]


async def view_attendance_requests_by_student(
    session: AsyncSession, student_id: str
) -> List[AttendanceRequestResponse]:
    stmt = select(AttendanceRequest).where(AttendanceRequest.student_id == student_id)
    rows = (await session.execute(stmt)).scalars().all()
    return [AttendanceRequestResponse.model_validate(r) for r in rows]


async def update_attendance_request(
    session: AsyncSession, request_id: int, updated: Dict[str, Any]
) -> AttendanceRequestResponse:
    req = await session.get(AttendanceRequest, request_id)
    if not req:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance request not found",
        )
    for key, value in updated.items():
        if hasattr(req, key):
            setattr(req, key, value)
    await session.commit()
    await session.refresh(req)
    return AttendanceRequestResponse.model_validate(req)


async def delete_attendance_request(session: AsyncSession, request_id: int) -> None:
    req = await session.get(AttendanceRequest, request_id)
    if not req:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance request not found",
        )
    await session.delete(req)
    await session.commit()


async def get_request_by_status(
    session: AsyncSession,
    request_status: Optional[RequestStatus] = None,
    module_id: Optional[str] = None,
) -> List[AttendanceRequestResponse]:
    stmt = select(AttendanceRequest)
    if request_status:
        stmt = stmt.where(AttendanceRequest.request_status == request_status)
    if module_id:
        stmt = stmt.where(AttendanceRequest.module_id == module_id)
    stmt = stmt.order_by(AttendanceRequest.created_at.desc())
    rows = (await session.execute(stmt)).scalars().all()
    return [AttendanceRequestResponse.model_validate(r) for r in rows]



async def _get_student_email(session: AsyncSession, student_id: str):
    result = await session.execute(select(Iam.email).where(Iam.iam_id == student_id))
    email = result.scalars().first()
    if not email:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    return email or ""


async def request_correction(
    session: AsyncSession,
    background_tasks: BackgroundTasks,
    data: AttendanceRequestCreate,
) -> AttendanceRequestResponse:
    attendance = await session.get(Attendance, data.attendance_id)
    if not attendance:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance record not found",
        )

    pending_stmt = select(AttendanceRequest).where(
        AttendanceRequest.attendance_id == data.attendance_id,
        AttendanceRequest.student_id == data.student_id,
        AttendanceRequest.request_status == RequestStatus.PENDING,
    )
    if (await session.execute(pending_stmt)).scalars().first():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="There is already a pending correction request for this attendance record",
        )

    new_req = AttendanceRequest(
        **data.model_dump(),
        request_status=RequestStatus.PENDING,
    )
    session.add(new_req)
    await session.commit()
    await session.refresh(new_req)

    student_email = await _get_student_email(session, data.student_id)
    if student_email:
        background_tasks.add_task(
            send_attendance_request_confirmation,
            student_email=student_email,
            request_data={
                "module_id": data.module_id,
                "attendance_date": attendance.created_at,
                "current_status": attendance.attendance_status,
                "proposed_status": data.proposed_status,
                "reason": data.reason,
            },
            request_id=new_req.request_id,
        )
    else:
        logger.warning(
            "Could not find email for student %s - confirmation email not sent",
            data.student_id,
        )

    background_tasks.add_task(
        create_new_request_notification,
        session,
        new_req.request_id,
        data.student_id,
    )
    return AttendanceRequestResponse.model_validate(new_req)


async def handle_correction(
    session: AsyncSession,
    background_tasks: BackgroundTasks,
    request_id: int,
    approval: CorrectionApproval,
) -> Dict[str, Any]:
    req = await session.get(AttendanceRequest, request_id)
    if not req:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance request not found",
        )
    if req.request_status != RequestStatus.PENDING:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This request has already been processed",
        )

    is_approved = approval.approved_status == req.proposed_status
    new_status = RequestStatus.APPROVED if is_approved else RequestStatus.REJECTED

    original_attendance = await session.get(Attendance, req.attendance_id)
    original_status = original_attendance.attendance_status if original_attendance else None

    now = datetime.now()
    req.request_status = new_status
    req.approved_status = approval.approved_status
    req.processed_by = approval.processed_by
    req.processed_at = now

    if is_approved and original_attendance:
        original_attendance.attendance_status = approval.approved_status
        original_attendance.updated_at = now
        logger.info(
            "Correction request %s approved and attendance updated to %s",
            request_id,
            approval.approved_status,
        )
    else:
        logger.info(
            "Correction request %s rejected. Proposed=%s admin_approved=%s",
            request_id,
            req.proposed_status,
            approval.approved_status,
        )

    await session.commit()

    student_email = await _get_student_email(session, req.student_id)
    if student_email:
        background_tasks.add_task(
            send_attendance_correction_notification,
            student_email=student_email,
            request_data={
                "module_id": req.module_id,
                "original_status": original_status,
                "proposed_status": req.proposed_status,
                "approved_status": approval.approved_status,
                "reason": req.reason,
            },
            is_approved=is_approved,
            processed_by=approval.processed_by,
        )
    else:
        logger.warning(
            "Could not find email for student %s - correction email not sent",
            req.student_id,
        )

    background_tasks.add_task(
        create_request_processed_notification,
        session,
        request_id,
        approval.processed_by,
        is_approved,
    )

    return {
        "status": new_status.value,
        "message": "Correction Approved" if is_approved else "Correction Rejected",
        "proposed_status": req.proposed_status,
        "approved_status": approval.approved_status,
        "is_approved": is_approved,
    }


# --------------------------------------------------------------------------- #
# Attendance rate / exam eligibility
# --------------------------------------------------------------------------- #


async def calculate_attendance_rate(
    session: AsyncSession, student_id: str, module_id: str
) -> float:
    stmt = select(Attendance).where(
        Attendance.student_id == student_id,
        Attendance.module_id == module_id,
    )
    records = (await session.execute(stmt)).scalars().all()
    if not records:
        return 0.0
    present = sum(1 for r in records if r.attendance_status in PRESENT_STATUSES)
    return round(present / len(records) * 100, 2)


async def check_exam_eligibility(
    session: AsyncSession, student_id: str, module_id: str
) -> Dict[str, Any]:
    reg_stmt = select(ModuleRegistration).where(
        ModuleRegistration.student_id == student_id,
        ModuleRegistration.module_id == module_id,
    )
    if not (await session.execute(reg_stmt)).scalars().first():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Student {student_id} is not registered for module {module_id}",
        )

    rate = await calculate_attendance_rate(session, student_id, module_id)
    is_eligible = rate >= ELIGIBILITY_THRESHOLD

    exam_stmt = select(Exam).where(
        Exam.student_id == student_id,
        Exam.module_id == module_id,
    )
    exam_record = (await session.execute(exam_stmt)).scalars().first()

    return {
        "attendance_rate": rate,
        "is_eligible": is_eligible,
        "exam_record": exam_record
        or {
            "student_id": student_id,
            "module_id": module_id,
            "attendance_rate": rate,
            "is_eligible": is_eligible,
            "_calculated": True,
        },
    }


async def update_eligibility_for_module(session: AsyncSession, module_id: str):
    reg_stmt = select(ModuleRegistration).where(ModuleRegistration.module_id == module_id)
    registrations = (await session.execute(reg_stmt)).scalars().all()
    results = []
    for reg in registrations:
        rate = await calculate_attendance_rate(session, reg.student_id, module_id)
        is_eligible = rate >= 80.0

        exam = await session.get(Exam, {"student_id": reg.student_id, "module_id": module_id})
        if exam:
            exam.attendance_rate = rate
            exam.is_eligible = is_eligible
        else:
            exam = Exam(student_id=reg.student_id, module_id=module_id, attendance_rate=rate, is_eligible=is_eligible)
            session.add(exam)
        results.append({"student_id": reg.student_id, "rate": rate, "is_eligible": is_eligible})
    await session.commit()
    return results


def _summarise_module_results(
    module_id: str, module_results: List[Dict[str, Any]]
) -> Dict[str, Any]:
    eligible = sum(
        1 for r in module_results if not r.get("error") and r.get("is_eligible") is True
    )
    ineligible = sum(
        1 for r in module_results if not r.get("error") and r.get("is_eligible") is False
    )
    errors = sum(1 for r in module_results if r.get("error"))
    processed = eligible + ineligible
    rate_str = f"{(eligible / processed * 100):.2f}%" if processed > 0 else "0%"
    return {
        "module_id": module_id,
        "students_processed": len(module_results),
        "students_success": eligible,
        "students_failed": ineligible,
        "students_with_errors": errors,
        "success_attendance_rate": rate_str,
        "results": module_results,
    }


def _aggregate_overview(
    success_count: int, failure_count: int, total_modules: int
) -> Dict[str, Any]:
    total_evaluated = success_count + failure_count
    overall_rate = (
        f"{(success_count / total_evaluated * 100):.2f}%" if total_evaluated > 0 else "0%"
    )
    return {
        "total_modules_processed": total_modules,
        "total_students_success": success_count,
        "total_students_failed": failure_count,
        "success_rate": overall_rate,
    }


async def update_exam_eligibility_for_all_modules(
    session: AsyncSession,
) -> Dict[str, Any]:
    module_ids = (
        await session.execute(select(distinct(ModuleRegistration.module_id)))
    ).scalars().all()
    if not module_ids:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No modules found with registrations",
        )

    all_results: List[Dict[str, Any]] = []
    success_count = 0
    failure_count = 0
    for module_id in module_ids:
        try:
            module_results = await update_eligibility_for_module(session, module_id)
            summary = _summarise_module_results(module_id, module_results)
            success_count += summary["students_success"]
            failure_count += summary["students_failed"]
            all_results.append(summary)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Error processing module %s", module_id)
            all_results.append(
                {
                    "module_id": module_id,
                    "error": str(exc),
                    "students_processed": 0,
                    "students_success": 0,
                    "students_failed": 0,
                    "students_with_errors": 1,
                    "success_attendance_rate": "0%",
                }
            )

    return {
        "summary": _aggregate_overview(success_count, failure_count, len(module_ids)),
        "modules": all_results,
    }


async def update_exam_eligibility_for_lecturer_modules(
    session: AsyncSession, lecturer_id: str
) -> Dict[str, Any]:
    module_ids = (
        await session.execute(
            select(distinct(ModuleRegistration.module_id)).where(
                ModuleRegistration.lecturer_id == lecturer_id
            )
        )
    ).scalars().all()
    if not module_ids:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No modules found for lecturer {lecturer_id}",
        )

    all_results: List[Dict[str, Any]] = []
    success_count = 0
    failure_count = 0
    for module_id in module_ids:
        try:
            module_results = await update_eligibility_for_module(session, module_id)
            summary = _summarise_module_results(module_id, module_results)
            success_count += summary["students_success"]
            failure_count += summary["students_failed"]
            all_results.append(summary)
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                "Error processing module %s for lecturer %s", module_id, lecturer_id
            )
            all_results.append(
                {
                    "module_id": module_id,
                    "error": str(exc),
                    "students_processed": 0,
                    "students_success": 0,
                    "students_failed": 0,
                    "students_with_errors": 1,
                    "success_attendance_rate": "0%",
                }
            )

    return {
        "lecturer_id": lecturer_id,
        "summary": _aggregate_overview(success_count, failure_count, len(module_ids)),
        "modules": all_results,
    }


async def get_exam_eligibility_by_module(
    session: AsyncSession, module_id: str
) -> List[Exam]:
    stmt = select(Exam).where(Exam.module_id == module_id).order_by(Exam.student_id)
    return (await session.execute(stmt)).scalars().all()


async def get_exam_eligibility_by_student(
    session: AsyncSession, student_id: str
) -> List[Exam]:
    stmt = select(Exam).where(Exam.student_id == student_id).order_by(Exam.module_id)
    return (await session.execute(stmt)).scalars().all()


# --------------------------------------------------------------------------- #
# Excel import / export
# --------------------------------------------------------------------------- #

async def import_attendance_from_excel(session: AsyncSession, file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    successful = []
    failed = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        student_id = str(row[0]).strip()
        module_id = str(row[1]).strip()
        status = str(row[2]).strip()

        reg_stmt = select(ModuleRegistration).where(
            ModuleRegistration.student_id == student_id,
            ModuleRegistration.module_id == module_id
        )
        if not (await session.execute(reg_stmt)).scalars().first():
            failed.append({"error": "Student not registered"})
            continue


        new_attendance = Attendance(
            student_id=student_id,
            module_id=module_id,
            attendance_status=status,
        )
        session.add(new_attendance)
        successful.append({"student_id": student_id, "module_id": module_id, "status": status})

    await session.commit()
    return {"successful": successful, "failed": failed}


status_colors = {
            "PRESENT": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "ABSENT": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "LATE": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "EXCUSED": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        }



async def export_eligibility_to_excel(
    module_id: str, eligibility_data: List[Dict[str, Any]]
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{module_id}_Eligibility"[:31]

    headers = ["Module ID", "Student ID", "Attendance Rate %", "Is Eligible"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for entry in eligibility_data:
        eligible = bool(entry.get("is_eligible"))
        ws.append(
            [
                module_id,
                entry.get("student_id"),
                entry.get("attendance_rate"),
                "YES" if eligible else "NO",
            ]
        )
        ws.cell(row=ws.max_row, column=4).fill = (
            status_colors["ELIGIBLE"] if eligible else status_colors["INELIGIBLE"]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _resolve_export_dir() -> Path:
    export_dir = Path(os.getcwd()) / EXPORT_DIR_NAME
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


async def _build_module_eligibility_workbook(
    session: AsyncSession,
    module_data: Dict[str, Any],
    lecturer_id: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    if module_data.get("error") or not module_data.get("results"):
        return None

    module_id = module_data["module_id"]
    info_stmt = (
        select(Module, Lecturer)
        .join(Lecturer, Lecturer.lecturer_id == Module.lecturer_id)
        .where(Module.module_id == module_id)
    )
    if lecturer_id:
        info_stmt = info_stmt.where(Module.lecturer_id == lecturer_id)
    info_row = (await session.execute(info_stmt)).first()
    if not info_row:
        logger.info("No module info found for module %s", module_id)
        return None
    module, lecturer = info_row

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{module_id}_Eligibility"[:31]

    columns = [
        ("Module ID", "module_id", 15),
        ("Module Name", "module_name", 30),
        ("Lecturer ID", "lecturer_id", 15),
        ("Student ID", "student_id", 15),
        ("Student Name", "student_name", 25),
        ("Attendance Rate", "attendance_rate", 18),
        ("Is Eligible", "is_eligible", 12),
    ]
    worksheet.append([col[0] for col in columns])
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for idx, (_, _, width) in enumerate(columns, start=1):
        column_letter = worksheet.cell(row=1, column=idx).column_letter
        worksheet.column_dimensions[column_letter].width = width

    student_rows: List[Dict[str, Any]] = []
    for entry in module_data["results"]:
        if entry.get("error"):
            continue
        exam_record = entry.get("exam_record")
        student_id = getattr(exam_record, "student_id", None) if exam_record else entry.get("student_id")
        if not student_id:
            continue
        student = await session.get(Student, student_id)
        if not student:
            logger.info("Student not found: %s", student_id)
            continue
        student_rows.append(
            {
                "module_id": module_id,
                "module_name": module.name,
                "lecturer_id": lecturer.lecturer_id,
                "student_id": student.student_id,
                "student_name": student.name,
                "attendance_rate": f"{float(entry['attendance_rate']):.2f}%",
                "is_eligible": "YES" if entry.get("is_eligible") else "NO",
            }
        )

    if not student_rows:
        return None

    for entry in student_rows:
        worksheet.append([entry[key] for _, key, _ in columns])
        cell = worksheet.cell(row=worksheet.max_row, column=len(columns))
        cell.fill = (
            status_colors["ELIGIBLE"]
            if entry["is_eligible"] == "YES"
            else status_colors["INELIGIBLE"]
        )

    return {
        "workbook": workbook,
        "students": student_rows,
        "module": module,
        "lecturer": lecturer,
    }


async def export_exam_eligibility_to_excel(session: AsyncSession) -> Dict[str, Any]:
    eligibility_data = await update_exam_eligibility_for_all_modules(session)
    export_dir = _resolve_export_dir()

    export_results: List[Dict[str, Any]] = []
    for module_data in eligibility_data["modules"]:
        module_id = module_data.get("module_id")
        try:
            if module_data.get("error"):
                export_results.append(
                    {
                        "module_id": module_id,
                        "error": f"Module processing error: {module_data['error']}",
                        "fileName": None,
                        "filePath": None,
                    }
                )
                continue

            built = await _build_module_eligibility_workbook(session, module_data)
            if not built:
                continue

            file_name = f"{module_id}_eligibility_for_exam.xlsx"
            file_path = export_dir / file_name
            built["workbook"].save(file_path)
            students = built["students"]
            export_results.append(
                {
                    "module_id": module_id,
                    "fileName": file_name,
                    "filePath": str(file_path),
                    "studentsProcessed": len(students),
                    "studentsEligible": sum(1 for s in students if s["is_eligible"] == "YES"),
                    "studentsIneligible": sum(1 for s in students if s["is_eligible"] == "NO"),
                    "errors": 0,
                }
            )
        except Exception as exc:  # noqa: BLE001
            logger.exception("Error exporting module %s", module_id)
            export_results.append(
                {
                    "module_id": module_id,
                    "error": str(exc),
                    "fileName": None,
                    "filePath": None,
                }
            )

    return {
        "exportDirectory": str(export_dir),
        "totalModules": len(eligibility_data["modules"]),
        "successfulExports": sum(1 for r in export_results if not r.get("error")),
        "failedExports": sum(1 for r in export_results if r.get("error")),
        "results": export_results,
    }


async def export_exam_eligibility_for_lecturer_to_excel(
    session: AsyncSession, lecturer_id: str
) -> Dict[str, Any]:
    eligibility_data = await update_exam_eligibility_for_lecturer_modules(session, lecturer_id)
    export_dir = _resolve_export_dir()

    export_results: List[Dict[str, Any]] = []
    for module_data in eligibility_data["modules"]:
        module_id = module_data.get("module_id")
        try:
            if module_data.get("error"):
                export_results.append(
                    {
                        "module_id": module_id,
                        "lecturer_id": lecturer_id,
                        "error": f"Module processing error: {module_data['error']}",
                        "fileName": None,
                        "filePath": None,
                    }
                )
                continue

            built = await _build_module_eligibility_workbook(session, module_data, lecturer_id)
            if not built:
                continue

            file_name = f"{lecturer_id}_{module_id}_eligibility_for_exam.xlsx"
            file_path = export_dir / file_name
            built["workbook"].save(file_path)
            students = built["students"]
            export_results.append(
                {
                    "module_id": module_id,
                    "lecturer_id": lecturer_id,
                    "fileName": file_name,
                    "filePath": str(file_path),
                    "studentsProcessed": len(students),
                    "studentsEligible": sum(1 for s in students if s["is_eligible"] == "YES"),
                    "studentsIneligible": sum(1 for s in students if s["is_eligible"] == "NO"),
                    "errors": 0,
                }
            )
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                "Error exporting module %s for lecturer %s", module_id, lecturer_id
            )
            export_results.append(
                {
                    "module_id": module_id,
                    "lecturer_id": lecturer_id,
                    "error": str(exc),
                    "fileName": None,
                    "filePath": None,
                }
            )

    return {
        "lecturer_id": lecturer_id,
        "exportDirectory": str(export_dir),
        "totalModules": len(eligibility_data["modules"]),
        "successfulExports": sum(1 for r in export_results if not r.get("error")),
        "failedExports": sum(1 for r in export_results if r.get("error")),
        "results": export_results,
    }





