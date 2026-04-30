import io
from typing import List, Optional

import openpyxl
from fastapi import HTTPException, status
from sqlalchemy import select, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models import Module, Lecturer, Program, Intake, Semester
from backend.schemas.module import ModuleResponse, ModuleCreate, ModuleUpdate


async def _ensure_fk_exists(session: AsyncSession, data: dict) -> None:
    if "lecturer_id" in data and data["lecturer_id"]:
        if not await session.get(Lecturer, data["lecturer_id"]):
            raise HTTPException(status_code=404, detail="Lecturer not found")

    if "program_id" in data and data["program_id"]:
        if not await session.get(Program, data["program_id"]):
            raise HTTPException(status_code=404, detail="Program not found")

    if "intake" in data and data["intake"] is not None:
        if not await session.get(Intake, data["intake"]):
            raise HTTPException(status_code=404, detail="Intake not found")

    if "semester_id" in data and data["semester_id"]:
        if not await session.get(Semester, data["semester_id"]):
            raise HTTPException(status_code=404, detail="Semester not found")


async def create_module(session: AsyncSession, data: ModuleCreate) -> ModuleResponse:
    existing_id = await session.get(Module, data.module_id)
    if existing_id:
        raise HTTPException(status_code=409, detail="Module ID already exists")
    if await check_module_exists(session, data.name, data.lecturer_id):
        raise HTTPException(
            status_code=409,
            detail="Module already exists with this name and lecturer",
        )
    await _ensure_fk_exists(session, data.model_dump())
    module = Module(**data.model_dump())
    session.add(module)
    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(
            status_code=409,
            detail="Could not create module due to duplicate or invalid data",
        )
    await session.refresh(module)
    return ModuleResponse.model_validate(module)


async def view_modules(session: AsyncSession) -> List[ModuleResponse]:
    result = await session.execute(select(Module))
    modules = result.scalars().all()
    return [ModuleResponse.model_validate(module) for module in modules]


async def check_module_exists(
    session: AsyncSession,
    name: str,
    lecturer_id: str,
    exclude_module_id: Optional[str] = None,
) -> bool:

    stmt = select(Module).where(
        Module.name == name,
        Module.lecturer_id == lecturer_id,
    )

    if exclude_module_id:
        stmt = stmt.where(Module.module_id != exclude_module_id)
    result = await session.execute(stmt)
    return result.scalars().first() is not None


async def update_module(
    session: AsyncSession,
    module_id: str,
    data: ModuleUpdate,
) -> ModuleResponse:
    module = await session.get(Module, module_id)

    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    update_data = data.model_dump(exclude_unset=True)

    if not update_data:
        return ModuleResponse.model_validate(module)
    await _ensure_fk_exists(session, update_data)
    next_name = update_data.get("name", module.name)
    next_lecturer_id = update_data.get("lecturer_id", module.lecturer_id)

    if await check_module_exists(
        session,
        next_name,
        next_lecturer_id,
        exclude_module_id=module_id,
    ):
        raise HTTPException(
            status_code=409,
            detail="Module already exists with this name and lecturer",
        )
    for field, value in update_data.items():
        setattr(module, field, value)

    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(
            status_code=409,
            detail="Could not update module due to invalid data",
        )
    await session.refresh(module)
    return ModuleResponse.model_validate(module)


async def delete_module(session: AsyncSession, module_id: str) -> None:
    module = await session.get(Module, module_id)
    if not module:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Module not found")
    await session.delete(module)
    await session.commit()


async def _process_excel(session: AsyncSession, file_contents: bytes):
    workbook = openpyxl.load_workbook(io.BytesIO(file_contents))
    sheet = workbook.active

    successful = []
    failed = []
    seen_module_ids = set()

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        module_id = str(row[0]).strip() if len(row) > 0 and row[0] else None
        name = str(row[1]).strip() if len(row) > 1 and row[1] else None
        lecturer_id = str(row[2]).strip() if len(row) > 2 and row[2] else None
        program_id = str(row[3]).strip() if len(row) > 3 and row[3] else None
        intake_raw = str(row[4]).strip() if len(row) > 4 and row[4] else None
        semester_id = str(row[5]).strip() if len(row) > 5 and row[5] else None
        camera_path = str(row[6]).strip() if len(row) > 6 and row[6] else None

        if not module_id or not name or not lecturer_id or not program_id or not intake_raw or not semester_id:
            failed.append({
                "row": row_index,
                "module_id": module_id,
                "name": name,
                "lecturer_id": lecturer_id,
                "program_id": program_id,
                "intake": intake_raw,
                "semester_id": semester_id,
                "error": "Missing required fields",
            })
            continue

        # Check duplicates
        try:
            intake = int(intake_raw)
        except ValueError:
            failed.append({
                "row": row_index,
                "module_id": module_id,
                "name": name,
                "error": f"Invalid intake value: {intake_raw}",
            })
            continue

        if module_id in seen_module_ids:
            failed.append({
                "row": row_index,
                "module_id": module_id,
                "name": name,
                "error": "Duplicate module ID in Excel file",
            })
            continue

        seen_module_ids.add(module_id)
        if await session.get(Module, module_id):
            failed.append({
                "row": row_index,
                "module_id": module_id,
                "name": name,
                "error": "Module ID already exists",
            })
            continue

        if await check_module_exists(session, name, lecturer_id):
            failed.append({
                "row": row_index,
                "module_id": module_id,
                "name": name,
                "lecturer_id": lecturer_id,
                "error": "Module with this name and lecturer already exists",
            })
            continue

        module_data = {
            "module_id": module_id,
            "name": name,
            "lecturer_id": lecturer_id,
            "program_id": program_id,
            "intake": intake,
            "semester_id": semester_id,
            "camera_path": camera_path,
        }
        try:
            await _ensure_fk_exists(session, module_data)
        except HTTPException as exc:
            failed.append({
                "row": row_index,
                **module_data,
                "error": exc.detail,
            })
            continue
        session.add(Module(**module_data))
        successful.append(module_data)
    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(
            status_code=409,
            detail="Could not import modules due to duplicate or invalid data",
        )
    return {"successful": successful, "failed": failed}


async def create_multiple_module_from_excel(session: AsyncSession, file_contents: bytes):
    return await _process_excel(session, file_contents)


async def get_camera_path(session: AsyncSession, module_id: str) -> dict:
    module = await session.get(Module, module_id)
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    if not module.camera_path:
        raise HTTPException(status_code=404, detail="No camera path configured for this module")
    return {
        "moduleId": module.module_id,
        "moduleName": module.name,
        "cameraPath": module.camera_path,
    }


async def set_camera_path(
    session: AsyncSession,
    module_id: str,
    camera_path: str,
) -> dict:
    module = await session.get(Module, module_id)
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    if not camera_path:
        raise HTTPException(status_code=400, detail="Camera path is required")
    module.camera_path = camera_path
    await session.commit()
    await session.refresh(module)
    return {
        "moduleId": module.module_id,
        "moduleName": module.name,
        "cameraPath": module.camera_path,
        "message": "Camera path updated successfully",
    }


async def delete_multiple_modules_from_excel(
    session: AsyncSession,
    file_contents: bytes,
):
    workbook = openpyxl.load_workbook(io.BytesIO(file_contents))
    sheet = workbook.active
    successful = []
    failed = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = str(row[0]).strip() if len(row) > 0 and row[0] else None
        lecturer_id = str(row[1]).strip() if len(row) > 1 and row[1] else None
        program_id = str(row[2]).strip() if len(row) > 2 and row[2] else None
        intake_raw = str(row[3]).strip() if len(row) > 3 and row[3] else None
        semester_id = str(row[4]).strip() if len(row) > 4 and row[4] else None
        if not name or not lecturer_id:
            failed.append({
                "row": row_index,
                "name": name,
                "lecturer_id": lecturer_id,
                "error": "Missing name or lecturer_id",
            })
            continue
        stmt = select(Module).where(
            Module.name == name,
            Module.lecturer_id == lecturer_id,
        )
        if program_id:
            stmt = stmt.where(Module.program_id == program_id)
        if intake_raw:
            try:
                stmt = stmt.where(Module.intake == int(intake_raw))
            except ValueError:
                failed.append({
                    "row": row_index,
                    "name": name,
                    "lecturer_id": lecturer_id,
                    "error": f"Invalid intake value: {intake_raw}",
                })
                continue
        if semester_id:
            stmt = stmt.where(Module.semester_id == semester_id)
        modules = (await session.execute(stmt)).scalars().all()
        if not modules:
            failed.append({
                "row": row_index,
                "name": name,
                "lecturer_id": lecturer_id,
                "program_id": program_id,
                "intake": intake_raw,
                "semester_id": semester_id,
                "error": "No matching modules found",
            })
            continue
        for module in modules:
            successful.append({
                "module_id": module.module_id,
                "name": module.name,
                "lecturer_id": module.lecturer_id,
                "program_id": module.program_id,
                "intake": module.intake,
                "semester_id": module.semester_id,
            })
            await session.delete(module)
    await session.commit()
    return {"successful": successful, "failed": failed}
